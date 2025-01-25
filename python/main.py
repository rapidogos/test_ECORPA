import csv
from openpyxl import load_workbook

# Cargar el archivo Excel
wb = load_workbook(r"D:\Andres-2\DESCARGAS 2025\pvpapn-2021-03-18-anexo-referencias-mas-vendidas.xlsx")

# Seleccionar la hoja 1
sheet = wb.worksheets[1]

# Inicializar una lista para almacenar los datos
datos = []

# Obtener los valores de las columnas C, E, F e I a partir de la fila 9
for row in sheet.iter_rows(min_row=9):
    # Extraer los valores de las columnas C (3), E (5), F (6) e I (9)
    c_valor = row[2].value  # Columna C (índice 2)nombre producto
    e_valor = row[4].value  # Columna E (índice 4)marca
    f_valor = row[5].value  # Columna F (índice 5)cant.vendidad
    i_valor = row[8].value  # Columna I (índice 8)precio    

    # Solo agregar filas si hay un valor en la Columna F (índice 5)cant.vendidad
    if f_valor is not None:
        datos.append((c_valor, e_valor, f_valor, i_valor))

# Ordenar los datos por los valores de la columna F (índice 2), en orden descendente
datos_ordenados = sorted(datos, key=lambda x: x[2], reverse=True)

# Seleccionar solo los 10 primeros
top_10 = datos_ordenados[:10]

# Imprimir los 10 resultados con los valores de las columnas C, E, F e I
for fila in top_10:
    print(f"Nombre Producto: {fila[0]}, Marca: {fila[1]}, Cant.Vendidas: {fila[2]}, Precio: {fila[3]}")

# Calcular el total de la columnas para las 10 filas seleccionadas
total_Cv = sum(fila[2] for fila in top_10)
total_pr=sum(fila[3] for fila in top_10)

# Guardar los 10 resultados en un archivo CSV
with open("top_10_resultados.csv", mode="w", newline="") as file:
    writer = csv.writer(file)
    
    # Escribir encabezado
    writer.writerow(["Nombre Producto", "Marca", "Cantidades Vendidas", "Precio"])

    
    # Escribir los datos
    for fila in top_10:
        writer.writerow(fila)
    
    # Escribir el total
    writer.writerow(["Total Cantidades vendidas[" ,total_Cv, "] y total precio[",total_pr,"]"])

print("Los datos se han guardado correctamente en 'top_10_resultados.csv'.")


